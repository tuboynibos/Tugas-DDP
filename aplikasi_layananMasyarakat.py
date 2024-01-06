import ttkbootstrap as ttk
import tkinter
from ttkbootstrap.constants import *        
from ttkbootstrap.toast import ToastNotification
from openpyxl import load_workbook
from openpyxl import Workbook
from pathlib import Path


class appHospital(ttk.Frame):
    def __init__(self, master_window):
        super().__init__(master_window, padding=(70, 50))
        self.pack(fill=BOTH, expand=YES)
        self.namaUser = ttk.StringVar(value="")
        self.addresUser = ttk.StringVar(value="")
        self.doctor = ttk.StringVar(value="")
        self.oldUser = ttk.DoubleVar(value=0)
        self.colors = master_window.style.colors
        
        instruction_text = "Masukkan Data Dengan Tepat!"
        instruction = ttk.Label(self, text=instruction_text, width=50)
        instruction.pack(fill=X, pady=10)
        
        
        self.create_form("Full Name: ", self.namaUser)
        self.create_form("Addres: ", self.addresUser)
        self.create_form("Dokter yang menangani: ", self.doctor)
        self.age_input = self.create_form("Age: ", self.oldUser)
        
        
        # self.create_meter()   
        self.create_button()
    
    def create_form(self, label, variable):
        form_field_container = ttk.Frame(self)
        form_field_container.pack(fill=X, expand=YES, pady=15)
        
        form_field_label = ttk.Label(master=form_field_container, text=label, width=25)
        form_field_label.pack(side=LEFT, padx=12) 
        
        form_input = ttk.Entry(master=form_field_container, textvariable=variable)
        form_input.pack(side=LEFT, padx=12, fill=X, expand=YES) 
        
        
        return form_input
        
    def create_button(self):
        button_container = ttk.Frame(self)
        button_container.pack(fill=X, expand=YES, pady=(15, 10))
        
        cancel_btn = ttk.Button(
            master=button_container,
            text="Batal",
            command=self.on_cancel,
            bootstyle=DANGER,
            width=6,
        )
        cancel_btn.pack(side=RIGHT, padx=5)    
        
        submit_btn = ttk.Button(
            master=button_container,
            text="Kirim",
            command=self.on_submit,
            bootstyle=SUCCESS,
            width=6,
        )
        submit_btn.pack(side=RIGHT, padx=5)    
    
    # ketika submit di klik
    def on_submit(self):
        
        name = self.namaUser.get()
        address = self.addresUser.get()
        doctor = self.doctor.get()
        age = self.age_input.get()

        nomor_antrian = self.get_next_queue_number()

        toast_message = f"Data Berhasil Disimpan\nNomor Antrian: {nomor_antrian}"
        toast = ToastNotification(title="Submission Successful", message=toast_message, duration=3000)
        toast.show_toast()

        lokasi_file = Path("Data_user.xlsx").expanduser().resolve()

        try:
            workbook = load_workbook(lokasi_file)
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet['A1'] = "Nomor Antrian"
            sheet['B1'] = "Nama Lengkap"
            sheet['C1'] = "Alamat"
            sheet['D1'] = "Dokter penanggung jawab"
            sheet['E1'] = "Age"

        sheet = workbook.active

        # Ambil nomor antrian dari file Excel
        try:
            nomor_antrian = int(sheet['A1'].value)
        except (ValueError, TypeError):
            nomor_antrian = 1

        sheet.append([nomor_antrian, name, address, doctor, age])

        # Update nomor antrian di file Excel
        sheet['A1'] = nomor_antrian + 1
        workbook.save(lokasi_file)

    def get_next_queue_number(self):
        lokasi_file = Path("Data_user.xlsx").expanduser().resolve()

        try:
            workbook = load_workbook(lokasi_file)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet['A1'] = "Nomor Antrian"


        try:
            nomor_antrian = int(sheet['A1'].value)
        except (ValueError, TypeError):
            nomor_antrian = 1

        return nomor_antrian
                
    def on_cancel(self): 
        self.quit()
    
        
if __name__ == "__main__":
    app = ttk.Window("appHospital", "superhero", resizable=(False, False))
    appHospital(app)
    app.eval('tk::PlaceWindow . center')
    app.mainloop()
